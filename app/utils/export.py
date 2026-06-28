"""Export utilities for CSV, Excel, and PDF generation."""

import csv
import datetime
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)


def export_to_csv(
    headers: list[str], rows: list[list[Any]], title: str = ""
) -> io.BytesIO:
    """Generate a CSV file in memory.

    Args:
        headers: Column header names.
        rows: List of row data lists.
        title: Optional title (included as first row).

    Returns:
        BytesIO buffer containing the CSV data.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    if title:
        writer.writerow([title])
        writer.writerow([])

    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)

    output = io.BytesIO()
    output.write(buffer.getvalue().encode("utf-8-sig"))
    output.seek(0)
    return output


def export_to_excel(
    headers: list[str],
    rows: list[list[Any]],
    title: str = "",
    sheet_name: str = "Report",
) -> io.BytesIO:
    """Generate an Excel file in memory.

    Args:
        headers: Column header names.
        rows: List of row data lists.
        title: Optional title for the sheet.
        sheet_name: Name of the Excel sheet.

    Returns:
        BytesIO buffer containing the Excel file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    current_row = 1

    # Title
    if title:
        ws.cell(row=current_row, column=1, value=title).font = Font(
            size=14, bold=True
        )
        current_row += 2

    # Headers
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    current_row += 1

    # Data rows
    for row_data in rows:
        for col, value in enumerate(row_data, 1):
            ws.cell(row=current_row, column=col, value=value)
        current_row += 1

    # Auto-size columns
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, current_row)
        )
        ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"].width = (
            min(max_len + 2, 40)
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_to_pdf(
    headers: list[str],
    rows: list[list[Any]],
    title: str = "",
    subtitle: str = "",
) -> io.BytesIO:
    """Generate a PDF report in memory.

    Args:
        headers: Column header names.
        rows: List of row data lists.
        title: Report title.
        subtitle: Report subtitle.

    Returns:
        BytesIO buffer containing the PDF.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    buffer = io.BytesIO()
    page_size = landscape(A4) if len(headers) > 5 else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    if title:
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Title"],
            fontSize=18,
            spaceAfter=6,
        )
        elements.append(Paragraph(title, title_style))

    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))

    elements.append(Spacer(1, 12))

    # Table data
    table_data = [headers] + [[str(cell) for cell in row] for row in rows]

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F8FC")]),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("TOPPADDING", (0, 1), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]
        )
    )

    elements.append(table)

    # Footer with generation time
    elements.append(Spacer(1, 20))
    footer_text = f"Generated: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    elements.append(Paragraph(footer_text, styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)
    return buffer
